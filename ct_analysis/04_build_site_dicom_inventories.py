"""Build a scalable SCAPIS DICOM inventory and per-site Excel reports.

The scanner reads DICOM headers only (never pixel data), stores every header tag
in SQLite, summarizes each SeriesInstanceUID, classifies 2D/3D/4D acquisitions,
and writes one analysis workbook per site plus a global 3D/4D workbook.

Example (Windows):
    python 04_build_site_dicom_inventories.py \
      --root "Q:\\users\\leejo\\data\\scapis\\datahub" \
      --output "Q:\\users\\leejo\\data\\scapis\\dicom_inventory"
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sqlite3
import sys
import time
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Sequence

import pandas as pd
import pydicom
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydicom.dataelem import DataElement
from pydicom.dataset import Dataset
from pydicom.multival import MultiValue
from pydicom.sequence import Sequence as DicomSequence

SITE_FOLDER_RE = re.compile(
    r"^ct[_-]site[_-](?P<site>\d+)[_-](?P<protocol>casc|ccta)"
    r"(?:[_-](?P<batch>\d+))?$",
    re.IGNORECASE,
)
PHASE_PERCENT_RE = re.compile(r"(?<!\d)(?P<phase>\d{1,3})\s*%")
SCAN_OPTION_PHASE_RE = re.compile(
    r"(?:TP|BESTPH_D_P)0*(?P<phase>\d{2,3})PC", re.IGNORECASE
)
SCAN_OPTION_HEART_RATE_RE = re.compile(
    r"OSCRATEAVG0*(?P<rate>\d{2,3})BPM", re.IGNORECASE
)
EXCEL_MAX_ROWS = 1_048_576
EXCEL_DATA_ROWS = EXCEL_MAX_ROWS - 1

FILE_COLUMNS = [
    "file_path",
    "file_size",
    "modified_ns",
    "source_folder",
    "site_number",
    "source_protocol",
    "archive_batch",
    "series_key",
    "patient_id",
    "study_uid",
    "series_uid",
    "sop_instance_uid",
    "sop_class_uid",
    "study_date",
    "study_time",
    "study_description",
    "series_date",
    "series_time",
    "series_number",
    "series_description",
    "modality",
    "manufacturer",
    "manufacturer_model",
    "institution_name",
    "station_name",
    "protocol_name",
    "body_part",
    "rows",
    "columns",
    "number_of_frames",
    "instance_number",
    "image_position",
    "image_orientation",
    "slice_position",
    "pixel_spacing",
    "slice_thickness",
    "spacing_between_slices",
    "frame_of_reference_uid",
    "image_type",
    "acquisition_date",
    "acquisition_time",
    "acquisition_number",
    "kvp",
    "tube_current",
    "exposure_time",
    "exposure",
    "revolution_time",
    "single_collimation_width",
    "total_collimation_width",
    "table_speed",
    "table_feed_per_rotation",
    "spiral_pitch_factor",
    "ctdi_vol",
    "convolution_kernel",
    "filter_type",
    "reconstruction_algorithm",
    "reconstruction_diameter",
    "data_collection_diameter",
    "contrast_agent",
    "contrast_route",
    "contrast_volume",
    "heart_rate",
    "trigger_time",
    "temporal_position_identifier",
    "temporal_position_index",
    "number_of_temporal_positions",
    "temporal_resolution",
    "cardiac_number_of_images",
    "nominal_cardiac_phase_percent",
    "nominal_cardiac_trigger_delay",
    "scan_options",
    "transfer_syntax_uid",
    "folder_path",
    "all_metadata_json",
]

IMPORTANT_DICOM_FIELDS = {
    "PatientID": "patient_id",
    "StudyInstanceUID": "study_uid",
    "SeriesInstanceUID": "series_uid",
    "SOPInstanceUID": "sop_instance_uid",
    "SOPClassUID": "sop_class_uid",
    "StudyDate": "study_date",
    "StudyTime": "study_time",
    "StudyDescription": "study_description",
    "SeriesDate": "series_date",
    "SeriesTime": "series_time",
    "SeriesNumber": "series_number",
    "SeriesDescription": "series_description",
    "Modality": "modality",
    "Manufacturer": "manufacturer",
    "ManufacturerModelName": "manufacturer_model",
    "InstitutionName": "institution_name",
    "StationName": "station_name",
    "ProtocolName": "protocol_name",
    "BodyPartExamined": "body_part",
    "Rows": "rows",
    "Columns": "columns",
    "NumberOfFrames": "number_of_frames",
    "InstanceNumber": "instance_number",
    "ImagePositionPatient": "image_position",
    "ImageOrientationPatient": "image_orientation",
    "PixelSpacing": "pixel_spacing",
    "SliceThickness": "slice_thickness",
    "SpacingBetweenSlices": "spacing_between_slices",
    "FrameOfReferenceUID": "frame_of_reference_uid",
    "ImageType": "image_type",
    "AcquisitionDate": "acquisition_date",
    "AcquisitionTime": "acquisition_time",
    "AcquisitionNumber": "acquisition_number",
    "KVP": "kvp",
    "XRayTubeCurrent": "tube_current",
    "ExposureTime": "exposure_time",
    "Exposure": "exposure",
    "RevolutionTime": "revolution_time",
    "SingleCollimationWidth": "single_collimation_width",
    "TotalCollimationWidth": "total_collimation_width",
    "TableSpeed": "table_speed",
    "TableFeedPerRotation": "table_feed_per_rotation",
    "SpiralPitchFactor": "spiral_pitch_factor",
    "CTDIvol": "ctdi_vol",
    "ConvolutionKernel": "convolution_kernel",
    "FilterType": "filter_type",
    "ReconstructionAlgorithm": "reconstruction_algorithm",
    "ReconstructionDiameter": "reconstruction_diameter",
    "DataCollectionDiameter": "data_collection_diameter",
    "ContrastBolusAgent": "contrast_agent",
    "ContrastBolusRoute": "contrast_route",
    "ContrastBolusVolume": "contrast_volume",
    "HeartRate": "heart_rate",
    "TriggerTime": "trigger_time",
    "TemporalPositionIdentifier": "temporal_position_identifier",
    "TemporalPositionIndex": "temporal_position_index",
    "NumberOfTemporalPositions": "number_of_temporal_positions",
    "TemporalResolution": "temporal_resolution",
    "CardiacNumberOfImages": "cardiac_number_of_images",
    "NominalPercentageOfCardiacPhase": "nominal_cardiac_phase_percent",
    "NominalCardiacTriggerDelayTime": "nominal_cardiac_trigger_delay",
    "ScanOptions": "scan_options",
}

SERIES_SQL = """
SELECT
    series_key,
    MAX(NULLIF(site_number, '')) AS site_number,
    MAX(NULLIF(source_protocol, '')) AS source_protocol,
    MAX(NULLIF(source_folder, '')) AS source_folder,
    MAX(NULLIF(archive_batch, '')) AS archive_batch,
    MAX(NULLIF(patient_id, '')) AS patient_id,
    MAX(NULLIF(study_uid, '')) AS study_uid,
    MAX(NULLIF(series_uid, '')) AS series_uid,
    MAX(NULLIF(study_date, '')) AS study_date,
    MAX(NULLIF(study_description, '')) AS study_description,
    MAX(NULLIF(series_number, '')) AS series_number,
    MAX(NULLIF(series_description, '')) AS series_description,
    MAX(NULLIF(modality, '')) AS modality,
    MAX(NULLIF(manufacturer, '')) AS manufacturer,
    MAX(NULLIF(manufacturer_model, '')) AS manufacturer_model,
    MAX(NULLIF(institution_name, '')) AS institution_name,
    MAX(NULLIF(station_name, '')) AS station_name,
    MAX(NULLIF(protocol_name, '')) AS protocol_name,
    MAX(NULLIF(body_part, '')) AS body_part,
    MAX(CAST(NULLIF(rows, '') AS INTEGER)) AS rows,
    MAX(CAST(NULLIF(columns, '') AS INTEGER)) AS columns,
    COUNT(*) AS n_dicom_files,
    COUNT(DISTINCT NULLIF(sop_instance_uid, '')) AS n_unique_sop_instances,
    SUM(CASE WHEN CAST(NULLIF(number_of_frames, '') AS INTEGER) > 1
             THEN CAST(number_of_frames AS INTEGER) ELSE 1 END) AS n_images_or_frames,
    COUNT(DISTINCT NULLIF(ROUND(CAST(slice_position AS REAL), 3), '')) AS n_unique_slice_positions,
    MIN(CAST(NULLIF(instance_number, '') AS REAL)) AS first_instance_number,
    MAX(CAST(NULLIF(instance_number, '') AS REAL)) AS last_instance_number,
    MAX(NULLIF(pixel_spacing, '')) AS pixel_spacing,
    MAX(CAST(NULLIF(slice_thickness, '') AS REAL)) AS slice_thickness_mm,
    MAX(CAST(NULLIF(spacing_between_slices, '') AS REAL)) AS spacing_between_slices_mm,
    MAX(NULLIF(frame_of_reference_uid, '')) AS frame_of_reference_uid,
    MAX(NULLIF(image_type, '')) AS image_type,
    MIN(NULLIF(acquisition_time, '')) AS acquisition_time_first,
    MAX(NULLIF(acquisition_time, '')) AS acquisition_time_last,
    MAX(NULLIF(kvp, '')) AS kvp,
    MAX(NULLIF(tube_current, '')) AS tube_current_ma,
    MAX(NULLIF(exposure_time, '')) AS exposure_time_ms,
    MAX(NULLIF(exposure, '')) AS exposure_mas,
    MAX(NULLIF(revolution_time, '')) AS revolution_time_s,
    MAX(NULLIF(single_collimation_width, '')) AS single_collimation_width_mm,
    MAX(NULLIF(total_collimation_width, '')) AS total_collimation_width_mm,
    MAX(NULLIF(table_speed, '')) AS table_speed,
    MAX(NULLIF(table_feed_per_rotation, '')) AS table_feed_per_rotation,
    MAX(NULLIF(spiral_pitch_factor, '')) AS spiral_pitch_factor,
    MAX(NULLIF(ctdi_vol, '')) AS ctdi_vol_mgy,
    MAX(NULLIF(convolution_kernel, '')) AS convolution_kernel,
    MAX(NULLIF(filter_type, '')) AS filter_type,
    MAX(NULLIF(reconstruction_algorithm, '')) AS reconstruction_algorithm,
    MAX(NULLIF(reconstruction_diameter, '')) AS reconstruction_diameter_mm,
    MAX(NULLIF(data_collection_diameter, '')) AS data_collection_diameter_mm,
    MAX(NULLIF(contrast_agent, '')) AS contrast_agent,
    MAX(NULLIF(contrast_route, '')) AS contrast_route,
    MAX(NULLIF(contrast_volume, '')) AS contrast_volume_ml,
    MAX(NULLIF(heart_rate, '')) AS heart_rate_bpm,
    COUNT(DISTINCT NULLIF(temporal_position_index, '')) AS n_temporal_position_indices,
    COUNT(DISTINCT NULLIF(temporal_position_identifier, '')) AS n_temporal_position_identifiers,
    COUNT(DISTINCT NULLIF(trigger_time, '')) AS n_trigger_times,
    COUNT(DISTINCT NULLIF(nominal_cardiac_phase_percent, '')) AS n_cardiac_phase_percentages,
    MAX(CAST(NULLIF(number_of_temporal_positions, '') AS INTEGER)) AS reported_temporal_positions,
    MAX(CAST(NULLIF(cardiac_number_of_images, '') AS INTEGER)) AS cardiac_number_of_images,
    GROUP_CONCAT(DISTINCT NULLIF(temporal_position_index, '')) AS temporal_position_indices,
    GROUP_CONCAT(DISTINCT NULLIF(temporal_position_identifier, '')) AS temporal_position_identifiers,
    GROUP_CONCAT(DISTINCT NULLIF(trigger_time, '')) AS trigger_times_ms,
    GROUP_CONCAT(DISTINCT NULLIF(nominal_cardiac_phase_percent, '')) AS cardiac_phase_percentages,
    MAX(NULLIF(nominal_cardiac_trigger_delay, '')) AS nominal_cardiac_trigger_delay_ms,
    MAX(NULLIF(temporal_resolution, '')) AS temporal_resolution_ms,
    MAX(NULLIF(scan_options, '')) AS scan_options,
    MAX(NULLIF(transfer_syntax_uid, '')) AS transfer_syntax_uid,
    MIN(folder_path) AS series_folder,
    MIN(file_path) AS example_file
FROM dicom_files
GROUP BY series_key
"""


@dataclass(frozen=True)
class SourceInfo:
    source_folder: str
    site_number: str
    protocol: str
    archive_batch: str


@dataclass(frozen=True)
class ScanResult:
    kind: str
    path: str
    record: dict[str, object] | None = None
    error: str = ""
    source: SourceInfo | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scan SCAPIS DICOM headers and build per-site Excel inventories."
    )
    parser.add_argument(
        "--root",
        default=r"Q:\users\leejo\data\scapis\datahub",
        help="Root containing ct_site_<site>_<casc|ccta>_<batch> folders.",
    )
    parser.add_argument(
        "--output",
        default=r"Q:\users\leejo\data\scapis\dicom_inventory",
        help="Output folder for SQLite and Excel reports.",
    )
    parser.add_argument(
        "--workers", type=int, default=8, help="Concurrent header readers."
    )
    parser.add_argument(
        "--batch-size", type=int, default=250, help="SQLite commit batch size."
    )
    parser.add_argument(
        "--report-only",
        action="store_true",
        help="Skip scanning and rebuild Excel reports from the existing SQLite database.",
    )
    parser.add_argument(
        "--no-all-tags",
        action="store_true",
        help="Do not retain the complete recursive header JSON (smaller database).",
    )
    parser.add_argument(
        "--include-file-sheets",
        action="store_true",
        help="Add per-file sheets to site workbooks; large sites may create huge files.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Development aid: stop after this many filesystem files (0 means all).",
    )
    return parser.parse_args()


def display_path(path: Path) -> str:
    text = str(path)
    if text.startswith("\\\\?\\UNC\\"):
        return "\\\\" + text[8:]
    if text.startswith("\\\\?\\"):
        return text[4:]
    return text


def io_path(path: Path) -> Path:
    if os.name != "nt":
        return path
    absolute = os.path.abspath(display_path(path))
    if absolute.startswith("\\\\?\\"):
        return Path(absolute)
    if absolute.startswith("\\\\"):
        return Path("\\\\?\\UNC\\" + absolute[2:])
    return Path("\\\\?\\" + absolute)


def source_info(path: Path, root: Path) -> SourceInfo:
    display = Path(display_path(path))
    display_root = Path(display_path(root))
    try:
        parts = display.relative_to(display_root).parts
    except ValueError:
        parts = display.parts

    for part in parts:
        match = SITE_FOLDER_RE.match(part)
        if match:
            return SourceInfo(
                source_folder=part,
                site_number=match.group("site"),
                protocol=match.group("protocol").upper(),
                archive_batch=match.group("batch") or "",
            )
    return SourceInfo("unclassified", "unknown", "unknown", "")


def iter_files(root: Path, output: Path, limit: int = 0) -> Iterator[Path]:
    output_resolved = os.path.normcase(os.path.abspath(display_path(output)))
    yielded = 0
    for current, dirs, files in os.walk(io_path(root)):
        current_path = Path(current)
        dirs[:] = [
            name
            for name in dirs
            if os.path.normcase(os.path.abspath(display_path(current_path / name)))
            != output_resolved
        ]
        for filename in files:
            yield current_path / filename
            yielded += 1
            if limit and yielded >= limit:
                return


def element_value(value: object, depth: int = 0) -> object:
    if isinstance(value, bytes):
        return f"<binary {len(value)} bytes>"
    if isinstance(value, DicomSequence):
        if depth >= 6:
            return "<sequence depth limit>"
        return [dataset_to_metadata(item, depth + 1) for item in value]
    if isinstance(value, (MultiValue, list, tuple)):
        return [element_value(item, depth + 1) for item in value]
    if value is None:
        return ""
    text = str(value)
    return text if len(text) <= 20_000 else text[:20_000] + "<truncated>"


def metadata_key(element: DataElement) -> str:
    tag_text = f"{element.tag.group:04X}{element.tag.element:04X}"
    keyword = element.keyword or re.sub(r"\W+", "_", element.name).strip("_")
    return f"{tag_text}_{keyword or 'Unknown'}"


def dataset_to_metadata(dataset: Dataset, depth: int = 0) -> dict[str, object]:
    metadata: dict[str, object] = {}
    for element in dataset:
        if element.keyword == "PixelData":
            continue
        metadata[metadata_key(element)] = element_value(element.value, depth)
    return metadata


def dicom_text(dataset: Dataset, keyword: str) -> str:
    raw_value = dataset.get(keyword)
    if raw_value is None:
        return ""
    value = element_value(raw_value)
    if isinstance(value, list):
        return "\\".join(str(item) for item in value)
    return str(value)


def parse_float_values(text: str, expected: int) -> list[float] | None:
    if not text:
        return None
    pieces = re.split(r"[\\,\[\]'\" ]+", text.strip())
    values = [piece for piece in pieces if piece]
    if len(values) < expected:
        return None
    try:
        return [float(value) for value in values[:expected]]
    except ValueError:
        return None


def slice_position(dataset: Dataset) -> str:
    position = parse_float_values(dicom_text(dataset, "ImagePositionPatient"), 3)
    orientation = parse_float_values(dicom_text(dataset, "ImageOrientationPatient"), 6)
    if position is None:
        return ""
    if orientation is None:
        return f"{position[2]:.6f}"

    row = orientation[:3]
    column = orientation[3:]
    normal = [
        row[1] * column[2] - row[2] * column[1],
        row[2] * column[0] - row[0] * column[2],
        row[0] * column[1] - row[1] * column[0],
    ]
    projected = sum(position[index] * normal[index] for index in range(3))
    return f"{projected:.6f}"


def has_dicom_identity(dataset: Dataset, has_preamble: bool) -> bool:
    if has_preamble:
        return True
    study_uid = dicom_text(dataset, "StudyInstanceUID")
    series_uid = dicom_text(dataset, "SeriesInstanceUID")
    sop_class_uid = dicom_text(dataset, "SOPClassUID")
    modality = dicom_text(dataset, "Modality")
    return bool(
        sop_class_uid or (study_uid and series_uid) or (series_uid and modality)
    )


def scan_file(path: Path, root: Path, keep_all_tags: bool) -> ScanResult:
    source = source_info(path, root)
    try:
        stat = path.stat()
        with path.open("rb") as stream:
            prefix = stream.read(132)
        has_preamble = len(prefix) == 132 and prefix[128:132] == b"DICM"
        dataset = pydicom.dcmread(
            str(path),
            stop_before_pixels=True,
            force=True,
            defer_size=1024,
        )
        if not has_dicom_identity(dataset, has_preamble):
            return ScanResult("not_dicom", display_path(path), source=source)

        record: dict[str, object] = {column: "" for column in FILE_COLUMNS}
        record.update(
            {
                "file_path": display_path(path),
                "file_size": stat.st_size,
                "modified_ns": stat.st_mtime_ns,
                "source_folder": source.source_folder,
                "site_number": source.site_number,
                "source_protocol": source.protocol,
                "archive_batch": source.archive_batch,
                "folder_path": display_path(path.parent),
            }
        )
        for keyword, column in IMPORTANT_DICOM_FIELDS.items():
            record[column] = dicom_text(dataset, keyword)

        record["slice_position"] = slice_position(dataset)
        if dataset.file_meta is not None:
            transfer_syntax = dataset.file_meta.get("TransferSyntaxUID")
            if transfer_syntax is not None:
                record["transfer_syntax_uid"] = str(transfer_syntax)

        series_uid = str(record["series_uid"])
        if series_uid:
            record["series_key"] = series_uid
        else:
            fallback = "|".join(
                [
                    str(record["study_uid"]),
                    str(record["patient_id"]),
                    str(record["series_number"]),
                    display_path(path.parent),
                ]
            )
            record["series_key"] = f"NO_UID:{fallback}"

        if keep_all_tags:
            record["all_metadata_json"] = json.dumps(
                dataset_to_metadata(dataset),
                ensure_ascii=False,
                separators=(",", ":"),
            )
        return ScanResult("dicom", display_path(path), record=record, source=source)
    except Exception as error:
        return ScanResult("error", display_path(path), error=str(error), source=source)


def connect_database(database_path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(database_path)
    connection.execute("PRAGMA journal_mode=DELETE")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA temp_store=MEMORY")
    connection.execute(
        "CREATE TABLE IF NOT EXISTS dicom_files ("
        + ",".join(
            f'"{column}" {"INTEGER" if column in {"file_size", "modified_ns"} else "TEXT"}'
            for column in FILE_COLUMNS
        )
        + ", PRIMARY KEY(file_path))"
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS read_errors (
            file_path TEXT PRIMARY KEY,
            source_folder TEXT,
            site_number TEXT,
            source_protocol TEXT,
            error TEXT
        )
        """
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_dicom_series ON dicom_files(series_key)"
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_dicom_site ON dicom_files(site_number)"
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_dicom_study ON dicom_files(study_uid)"
    )
    connection.commit()
    return connection


def is_unchanged(connection: sqlite3.Connection, path: Path) -> bool:
    try:
        stat = path.stat()
    except OSError:
        return False
    row = connection.execute(
        "SELECT file_size, modified_ns FROM dicom_files WHERE file_path = ?",
        (display_path(path),),
    ).fetchone()
    return bool(row and row[0] == stat.st_size and row[1] == stat.st_mtime_ns)


def insert_records(
    connection: sqlite3.Connection, records: Sequence[dict[str, object]]
) -> None:
    if not records:
        return
    placeholders = ",".join("?" for _ in FILE_COLUMNS)
    columns = ",".join(f'"{column}"' for column in FILE_COLUMNS)
    connection.executemany(
        f"INSERT OR REPLACE INTO dicom_files ({columns}) VALUES ({placeholders})",
        [[record[column] for column in FILE_COLUMNS] for record in records],
    )


def insert_error(connection: sqlite3.Connection, result: ScanResult) -> None:
    source = result.source or SourceInfo("unclassified", "unknown", "unknown", "")
    connection.execute(
        """
        INSERT OR REPLACE INTO read_errors
        (file_path, source_folder, site_number, source_protocol, error)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            result.path,
            source.source_folder,
            source.site_number,
            source.protocol,
            result.error,
        ),
    )


def scan_root(
    root: Path,
    output: Path,
    connection: sqlite3.Connection,
    workers: int,
    batch_size: int,
    keep_all_tags: bool,
    limit: int,
) -> None:
    counts = {"filesystem": 0, "dicom": 0, "not_dicom": 0, "error": 0, "unchanged": 0}
    pending: set[Future[ScanResult]] = set()
    records: list[dict[str, object]] = []
    started = time.time()

    def consume(future: Future[ScanResult]) -> None:
        result = future.result()
        counts[result.kind] += 1
        if result.kind == "dicom" and result.record is not None:
            records.append(result.record)
        elif result.kind == "error":
            insert_error(connection, result)

    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        for path in iter_files(root, output, limit):
            counts["filesystem"] += 1
            if is_unchanged(connection, path):
                counts["unchanged"] += 1
                continue
            pending.add(executor.submit(scan_file, path, root, keep_all_tags))

            if len(pending) >= max(workers * 8, batch_size):
                done = next(as_completed(pending))
                pending.remove(done)
                consume(done)

            if len(records) >= batch_size:
                insert_records(connection, records)
                connection.commit()
                records.clear()

            if counts["filesystem"] % 5_000 == 0:
                elapsed = max(time.time() - started, 0.001)
                print(
                    f"Scanned {counts['filesystem']:,} filesystem files "
                    f"({counts['dicom']:,} new DICOM, "
                    f"{counts['unchanged']:,} unchanged, "
                    f"{counts['error']:,} errors) at "
                    f"{counts['filesystem'] / elapsed:.1f} files/s"
                )

        for future in as_completed(pending):
            consume(future)
            if len(records) >= batch_size:
                insert_records(connection, records)
                connection.commit()
                records.clear()

    insert_records(connection, records)
    connection.commit()
    elapsed = time.time() - started
    print("\nScan complete")
    print(f"  Files encountered: {counts['filesystem']:,}")
    print(f"  New/changed DICOM: {counts['dicom']:,}")
    print(f"  Unchanged DICOM:   {counts['unchanged']:,}")
    print(f"  Non-DICOM files:   {counts['not_dicom']:,}")
    print(f"  Read errors:       {counts['error']:,}")
    print(f"  Elapsed:           {elapsed / 60:.1f} minutes")


def numeric(value: object, default: float = 0.0) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return default


def series_role(row: pd.Series) -> str:
    description = " ".join(
        str(row.get(column, ""))
        for column in ["series_description", "protocol_name", "image_type"]
    ).lower()
    n_files = int(numeric(row.get("n_dicom_files")))
    source_protocol = str(row.get("source_protocol", "")).upper()
    thickness = numeric(row.get("slice_thickness_mm"), -1.0)

    if any(
        token in description
        for token in ["localizer", "topogram", "topo", "scout", "survey"]
    ):
        return "localizer"
    if "testbolus" in description or "test bolus" in description:
        return "test_bolus"
    if n_files <= 2:
        return "localizer_or_single_image"
    if source_protocol == "CASC":
        return "calcium_scoring"
    if source_protocol == "CCTA":
        return "ccta"
    if any(token in description for token in ["calcium", "casc", "b35f", "score"]):
        return "calcium_scoring"
    if any(token in description for token in ["ccta", "coronary", "angio", "cta"]):
        return "ccta"
    if thickness >= 2.5:
        return "probable_calcium_scoring"
    if 0 < thickness < 1.0 and n_files >= 50:
        return "probable_ccta"
    return "other"


def description_phase_percent(description: object) -> str:
    match = PHASE_PERCENT_RE.search(str(description))
    if not match:
        return ""
    value = int(match.group("phase"))
    return str(value) if 0 <= value <= 100 else ""


def scan_options_phase_percent(scan_options: object) -> str:
    match = SCAN_OPTION_PHASE_RE.search(str(scan_options))
    if not match:
        return ""
    value = int(match.group("phase"))
    return str(value) if 0 <= value <= 100 else ""


def scan_options_heart_rate(scan_options: object) -> str:
    match = SCAN_OPTION_HEART_RATE_RE.search(str(scan_options))
    return match.group("rate") if match else ""


def normalized_series_family(description: object) -> str:
    text = str(description).lower()
    text = PHASE_PERCENT_RE.sub("<phase>", text)
    text = re.sub(r"\b(best\s*diast|phase)\b", "", text)
    text = re.sub(r"\bno[_ -]?stack[_ -]?overlap\b", "", text)
    return re.sub(r"\W+", " ", text).strip()


def initial_volume_class(row: pd.Series) -> tuple[str, str, int]:
    role = str(row["series_role"])
    files = int(numeric(row.get("n_dicom_files")))
    frames = int(numeric(row.get("n_images_or_frames"), files))
    slices = int(numeric(row.get("n_unique_slice_positions")))
    temporal_indices = int(numeric(row.get("n_temporal_position_indices")))
    temporal_identifiers = int(numeric(row.get("n_temporal_position_identifiers")))
    cardiac_phases = int(numeric(row.get("n_cardiac_phase_percentages")))
    trigger_times = int(numeric(row.get("n_trigger_times")))
    reported_temporal = int(numeric(row.get("reported_temporal_positions")))
    cardiac_images = int(numeric(row.get("cardiac_number_of_images")))
    observed_temporal = max(temporal_indices, temporal_identifiers, cardiac_phases)

    if role.startswith("localizer") or (files <= 2 and frames <= 2):
        return "2D/localizer", "too few spatial images for a volume", 1

    has_spatial_volume = slices >= 3 or frames >= 3
    if has_spatial_volume and observed_temporal >= 2:
        return (
            "4D",
            "multiple temporal positions observed within one series",
            observed_temporal,
        )
    plausible_cardiac_phases = cardiac_images if 2 <= cardiac_images <= 100 else 0
    if has_spatial_volume and (
        reported_temporal >= 2 or plausible_cardiac_phases or trigger_times >= 2
    ):
        temporal = max(reported_temporal, plausible_cardiac_phases, trigger_times)
        return (
            "possible_4D",
            "temporal values are reported but phase-to-slice evidence needs review",
            temporal,
        )
    if has_spatial_volume:
        return "3D", "single spatial volume", 1
    return "2D/other", "insufficient spatial-position evidence", 1


def cardiac_sampling(row: pd.Series) -> str:
    volume_class = str(row.get("volume_class", ""))
    if volume_class.startswith("4D"):
        return "multi-phase (4D evidence)"
    if volume_class == "possible_4D":
        return "possible multi-phase; review"
    if str(row.get("phase_for_family", "")).strip():
        return "single reconstructed cardiac phase"
    if str(row.get("series_role", "")) == "ccta":
        return "no multi-phase evidence"
    return "not determined / not cardiac"


def normalized_phase(value: str) -> str:
    try:
        number = float(value)
    except ValueError:
        return value
    return f"{number:g}" if 0 <= number <= 100 else value


def classify_series(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return series

    classified = series.copy()
    classified["series_role"] = classified.apply(series_role, axis=1)
    classified["description_phase_percent"] = classified["series_description"].apply(
        description_phase_percent
    )
    classified["scan_options_phase_percent"] = classified["scan_options"].apply(
        scan_options_phase_percent
    )
    classified["scan_options_average_heart_rate"] = classified["scan_options"].apply(
        scan_options_heart_rate
    )
    classified["series_family_description"] = classified["series_description"].apply(
        normalized_series_family
    )

    initial = classified.apply(initial_volume_class, axis=1)
    classified["volume_class"] = [item[0] for item in initial]
    classified["volume_class_reason"] = [item[1] for item in initial]
    classified["temporal_positions_estimated"] = [item[2] for item in initial]

    cardiac_phase = classified["cardiac_phase_percentages"].fillna("").astype(str)
    description_phase = classified["description_phase_percent"].fillna("").astype(str)
    scan_options_phase = classified["scan_options_phase_percent"].fillna("").astype(str)
    classified["phase_for_family"] = cardiac_phase.where(
        cardiac_phase != "", description_phase
    )
    classified["phase_for_family"] = classified["phase_for_family"].where(
        classified["phase_for_family"] != "", scan_options_phase
    )
    dicom_heart_rate = classified["heart_rate_bpm"].fillna("").astype(str)
    classified["average_heart_rate_bpm"] = dicom_heart_rate.where(
        dicom_heart_rate != "", classified["scan_options_average_heart_rate"]
    )

    family_columns = [
        "patient_id",
        "study_uid",
        "series_family_description",
        "rows",
        "columns",
        "n_unique_slice_positions",
    ]
    for _, indices in classified.groupby(family_columns, dropna=False).groups.items():
        group = classified.loc[indices]
        phases: set[str] = set()
        for value in group["phase_for_family"]:
            phases.update(
                normalized_phase(part.strip())
                for part in str(value).split(",")
                if part.strip()
            )
        spatial = group["volume_class"].isin(["3D", "possible_4D", "4D"])
        if len(phases) >= 2 and spatial.sum() >= 2:
            classified.loc[indices, "volume_class"] = "4D (multi-series)"
            classified.loc[indices, "volume_class_reason"] = (
                "multiple cardiac phases reconstructed as separate 3D series"
            )
            classified.loc[indices, "temporal_positions_estimated"] = len(phases)

    classified["cardiac_sampling"] = classified.apply(cardiac_sampling, axis=1)
    classified["has_contrast"] = (
        classified["contrast_agent"].fillna("").astype(str).str.strip().ne("")
    )
    classified["duplicate_sop_count"] = (
        classified["n_dicom_files"].fillna(0).astype(int)
        - classified["n_unique_sop_instances"].fillna(0).astype(int)
    ).clip(lower=0)
    classified["data_quality_flags"] = classified.apply(series_quality_flags, axis=1)

    preferred_columns = [
        "site_number",
        "source_protocol",
        "source_folder",
        "archive_batch",
        "patient_id",
        "study_uid",
        "series_uid",
        "series_number",
        "series_description",
        "series_role",
        "volume_class",
        "volume_class_reason",
        "temporal_positions_estimated",
        "phase_for_family",
        "cardiac_sampling",
        "n_dicom_files",
        "n_images_or_frames",
        "n_unique_slice_positions",
        "rows",
        "columns",
        "pixel_spacing",
        "slice_thickness_mm",
        "spacing_between_slices_mm",
        "manufacturer",
        "manufacturer_model",
        "institution_name",
        "station_name",
        "protocol_name",
        "body_part",
        "kvp",
        "tube_current_ma",
        "exposure_time_ms",
        "exposure_mas",
        "revolution_time_s",
        "single_collimation_width_mm",
        "total_collimation_width_mm",
        "table_speed",
        "table_feed_per_rotation",
        "spiral_pitch_factor",
        "ctdi_vol_mgy",
        "convolution_kernel",
        "filter_type",
        "reconstruction_algorithm",
        "reconstruction_diameter_mm",
        "data_collection_diameter_mm",
        "has_contrast",
        "contrast_agent",
        "contrast_route",
        "contrast_volume_ml",
        "heart_rate_bpm",
        "average_heart_rate_bpm",
        "n_temporal_position_indices",
        "n_temporal_position_identifiers",
        "n_trigger_times",
        "n_cardiac_phase_percentages",
        "reported_temporal_positions",
        "cardiac_number_of_images",
        "temporal_position_indices",
        "temporal_position_identifiers",
        "trigger_times_ms",
        "cardiac_phase_percentages",
        "nominal_cardiac_trigger_delay_ms",
        "temporal_resolution_ms",
        "study_date",
        "study_description",
        "acquisition_time_first",
        "acquisition_time_last",
        "modality",
        "image_type",
        "scan_options",
        "frame_of_reference_uid",
        "transfer_syntax_uid",
        "duplicate_sop_count",
        "data_quality_flags",
        "series_folder",
        "example_file",
        "series_key",
    ]
    remaining = [
        column for column in classified.columns if column not in preferred_columns
    ]
    return classified[preferred_columns + remaining]


def series_quality_flags(row: pd.Series) -> str:
    flags: list[str] = []
    if not str(row.get("series_uid", "")).strip():
        flags.append("missing SeriesInstanceUID")
    if not str(row.get("study_uid", "")).strip():
        flags.append("missing StudyInstanceUID")
    if not str(row.get("patient_id", "")).strip():
        flags.append("missing PatientID")
    if not str(row.get("series_description", "")).strip():
        flags.append("missing SeriesDescription")
    role = str(row.get("series_role", ""))
    if not role.startswith("localizer"):
        if not str(row.get("slice_thickness_mm", "")).strip():
            flags.append("missing SliceThickness")
        if not str(row.get("pixel_spacing", "")).strip():
            flags.append("missing PixelSpacing")
    if numeric(row.get("duplicate_sop_count")) > 0:
        flags.append("duplicate SOPInstanceUID")
    if str(row.get("volume_class", "")) == "possible_4D":
        flags.append("temporal classification needs review")
    return "; ".join(flags)


def load_series(connection: sqlite3.Connection) -> pd.DataFrame:
    return classify_series(pd.read_sql_query(SERIES_SQL, connection))


def build_site_summary(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    grouped = series.groupby("site_number", dropna=False)
    return grouped.agg(
        source_folders=("source_folder", "nunique"),
        patients=("patient_id", "nunique"),
        studies=("study_uid", "nunique"),
        series=("series_key", "count"),
        dicom_files=("n_dicom_files", "sum"),
        casc_series=("source_protocol", lambda values: (values == "CASC").sum()),
        ccta_series=("source_protocol", lambda values: (values == "CCTA").sum()),
        volumes_3d=("volume_class", lambda values: (values == "3D").sum()),
        volumes_4d=(
            "volume_class",
            lambda values: values.astype(str).str.startswith("4D").sum(),
        ),
        possible_4d=("volume_class", lambda values: (values == "possible_4D").sum()),
        series_with_quality_flags=(
            "data_quality_flags",
            lambda values: values.astype(str).str.len().gt(0).sum(),
        ),
    ).reset_index()


def build_folder_index(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    return (
        series.groupby(
            ["site_number", "source_protocol", "source_folder", "archive_batch"],
            dropna=False,
        )
        .agg(
            patients=("patient_id", "nunique"),
            studies=("study_uid", "nunique"),
            series=("series_key", "count"),
            dicom_files=("n_dicom_files", "sum"),
            volumes_3d=("volume_class", lambda values: (values == "3D").sum()),
            volumes_4d=(
                "volume_class",
                lambda values: values.astype(str).str.startswith("4D").sum(),
            ),
            folder_path=("series_folder", lambda values: common_source_parent(values)),
        )
        .reset_index()
    )


def common_source_parent(values: Iterable[object]) -> str:
    paths = [str(value) for value in values if str(value)]
    if not paths:
        return ""
    source_folder = Path(paths[0])
    for parent in [source_folder, *source_folder.parents]:
        if parent.name and SITE_FOLDER_RE.match(parent.name):
            return str(parent)
    return str(source_folder)


def build_protocol_summary(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    working = series.copy()
    for column in [
        "slice_thickness_mm",
        "spacing_between_slices_mm",
        "kvp",
        "revolution_time_s",
        "spiral_pitch_factor",
    ]:
        working[column] = pd.to_numeric(working[column], errors="coerce").round(3)
    keys = [
        "site_number",
        "source_protocol",
        "series_role",
        "manufacturer",
        "manufacturer_model",
        "protocol_name",
        "slice_thickness_mm",
        "spacing_between_slices_mm",
        "pixel_spacing",
        "kvp",
        "convolution_kernel",
        "reconstruction_algorithm",
        "has_contrast",
        "revolution_time_s",
        "spiral_pitch_factor",
    ]
    return (
        working.groupby(keys, dropna=False)
        .agg(
            patients=("patient_id", "nunique"),
            studies=("study_uid", "nunique"),
            series=("series_key", "count"),
            dicom_files=("n_dicom_files", "sum"),
            min_slices=("n_unique_slice_positions", "min"),
            median_slices=("n_unique_slice_positions", "median"),
            max_slices=("n_unique_slice_positions", "max"),
            cardiac_phases=("phase_for_family", lambda values: distinct_join(values)),
        )
        .reset_index()
        .sort_values(
            ["site_number", "source_protocol", "series"], ascending=[True, True, False]
        )
    )


def distinct_join(values: Iterable[object]) -> str:
    distinct = sorted({str(value).strip() for value in values if str(value).strip()})
    return ", ".join(distinct)


def build_patient_summary(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    return (
        series.groupby(["site_number", "patient_id"], dropna=False)
        .agg(
            studies=("study_uid", "nunique"),
            series=("series_key", "count"),
            dicom_files=("n_dicom_files", "sum"),
            protocols=("source_protocol", distinct_join),
            series_roles=("series_role", distinct_join),
            volume_classes=("volume_class", distinct_join),
            first_study_date=("study_date", "min"),
            last_study_date=("study_date", "max"),
            example_folder=("series_folder", "first"),
        )
        .reset_index()
    )


def build_study_summary(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    return (
        series.groupby(["site_number", "patient_id", "study_uid"], dropna=False)
        .agg(
            study_date=("study_date", "first"),
            study_description=("study_description", "first"),
            protocols=("source_protocol", distinct_join),
            series=("series_key", "count"),
            dicom_files=("n_dicom_files", "sum"),
            series_roles=("series_role", distinct_join),
            volume_classes=("volume_class", distinct_join),
            scanner_models=("manufacturer_model", distinct_join),
            example_folder=("series_folder", "first"),
        )
        .reset_index()
    )


def build_quality_summary(series: pd.DataFrame) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame()
    flagged = series[series["data_quality_flags"].fillna("").astype(str).str.len() > 0]
    if flagged.empty:
        return pd.DataFrame([{"issue": "No series-level quality flags", "series": 0}])
    counts: dict[str, int] = {}
    for value in flagged["data_quality_flags"]:
        for flag in str(value).split("; "):
            counts[flag] = counts.get(flag, 0) + 1
    return pd.DataFrame(
        [{"issue": issue, "series": count} for issue, count in sorted(counts.items())]
    )


def windows_file_uri(value: object) -> str:
    text = str(value).strip()
    if not text:
        return ""
    normalized = text.replace("\\", "/")
    if re.match(r"^[A-Za-z]:/", normalized):
        return "file:///" + normalized.replace(" ", "%20")
    if normalized.startswith("//"):
        return "file:" + normalized.replace(" ", "%20")
    try:
        return Path(text).resolve().as_uri()
    except ValueError:
        return text


def write_dataframe_chunks(
    writer: pd.ExcelWriter,
    dataframe: pd.DataFrame,
    base_sheet_name: str,
) -> list[str]:
    if dataframe.empty:
        dataframe.to_excel(writer, sheet_name=base_sheet_name[:31], index=False)
        return [base_sheet_name[:31]]
    sheets: list[str] = []
    for index, start in enumerate(range(0, len(dataframe), EXCEL_DATA_ROWS), 1):
        suffix = f"_{index}" if len(dataframe) > EXCEL_DATA_ROWS else ""
        sheet_name = (base_sheet_name[: 31 - len(suffix)] + suffix)[:31]
        dataframe.iloc[start : start + EXCEL_DATA_ROWS].to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        sheets.append(sheet_name)
    return sheets


def format_workbook(path: Path, hyperlink_columns: set[str]) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        headers = {
            cell.column: str(cell.value)
            for cell in worksheet[1]
            if cell.value is not None
        }
        for column_index, header in headers.items():
            max_length = len(header)
            for row_index in range(2, min(worksheet.max_row, 5000) + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if cell.value is not None:
                    max_length = max(max_length, min(len(str(cell.value)), 80))
            if header in hyperlink_columns:
                for row_index in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_index, column=column_index)
                    if str(cell.value or "").strip():
                        cell.hyperlink = windows_file_uri(cell.value)
                        cell.style = "Hyperlink"
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10), 60
            )
    workbook.save(path)


def read_errors(
    connection: sqlite3.Connection, site_number: str | None = None
) -> pd.DataFrame:
    if site_number is None:
        return pd.read_sql_query(
            "SELECT * FROM read_errors ORDER BY site_number, source_folder, file_path",
            connection,
        )
    return pd.read_sql_query(
        "SELECT * FROM read_errors WHERE site_number = ? ORDER BY source_folder, file_path",
        connection,
        params=(site_number,),
    )


def file_inventory(connection: sqlite3.Connection, site_number: str) -> pd.DataFrame:
    selected = [column for column in FILE_COLUMNS if column != "all_metadata_json"]
    columns = ",".join(f'"{column}"' for column in selected)
    return pd.read_sql_query(
        f"SELECT {columns} FROM dicom_files WHERE site_number = ? ORDER BY source_folder, patient_id, study_uid, series_uid, instance_number",
        connection,
        params=(site_number,),
    )


def workbook_guide(database_path: Path) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "output": "Series",
                "contents": "One row per DICOM SeriesInstanceUID with acquisition, reconstruction, contrast, cardiac, and geometry metadata.",
            },
            {
                "output": "3D_4D_Images",
                "contents": "All confirmed 3D, confirmed 4D, and conservative possible-4D series with source-folder links.",
            },
            {
                "output": "Protocols",
                "contents": "Grouped protocol signatures for comparing acquisition and reconstruction settings.",
            },
            {
                "output": "Patients / Studies",
                "contents": "Patient- and study-level counts and links. PatientName is intentionally omitted from Excel.",
            },
            {
                "output": "SQLite database",
                "contents": f"{database_path} contains one row per DICOM file and all non-pixel header tags as JSON.",
            },
            {
                "output": "4D rule",
                "contents": "Confirmed only when multiple temporal positions occur in a series or distinct cardiac phases form a conservative multi-series family; reported-only counts are possible_4D.",
            },
        ]
    )


def create_site_workbook(
    site: str,
    site_series: pd.DataFrame,
    connection: sqlite3.Connection,
    output: Path,
    database_path: Path,
    include_file_sheets: bool,
) -> Path:
    workbook_path = output / f"site_{site}_dicom_inventory.xlsx"
    images = site_series[
        site_series["volume_class"].isin(
            ["3D", "4D", "4D (multi-series)", "possible_4D"]
        )
    ].copy()
    overview = build_site_summary(site_series)
    folders = build_folder_index(site_series)
    protocols = build_protocol_summary(site_series)
    patients = build_patient_summary(site_series)
    studies = build_study_summary(site_series)
    quality = build_quality_summary(site_series)
    errors = read_errors(connection, site)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        workbook_guide(database_path).to_excel(writer, sheet_name="Guide", index=False)
        overview.to_excel(writer, sheet_name="Overview", index=False)
        folders.to_excel(writer, sheet_name="Source_Folders", index=False)
        protocols.to_excel(writer, sheet_name="Protocols", index=False)
        patients.to_excel(writer, sheet_name="Patients", index=False)
        studies.to_excel(writer, sheet_name="Studies", index=False)
        write_dataframe_chunks(writer, site_series, "Series")
        write_dataframe_chunks(
            writer,
            site_series[site_series["source_protocol"] == "CASC"],
            "CASC_Series",
        )
        write_dataframe_chunks(
            writer,
            site_series[site_series["source_protocol"] == "CCTA"],
            "CCTA_Series",
        )
        write_dataframe_chunks(writer, images, "3D_4D_Images")
        quality.to_excel(writer, sheet_name="Quality_Summary", index=False)
        write_dataframe_chunks(writer, errors, "Read_Errors")
        if include_file_sheets:
            write_dataframe_chunks(
                writer, file_inventory(connection, site), "DICOM_Files"
            )

    format_workbook(
        workbook_path,
        {"folder_path", "series_folder", "example_file", "example_folder", "file_path"},
    )
    return workbook_path


def create_master_workbook(
    series: pd.DataFrame,
    connection: sqlite3.Connection,
    output: Path,
    database_path: Path,
) -> Path:
    workbook_path = output / "SCAPIS_all_sites_3D_4D_inventory.xlsx"
    images = series[
        series["volume_class"].isin(["3D", "4D", "4D (multi-series)", "possible_4D"])
    ].copy()
    site_summary = build_site_summary(series)
    folder_index = build_folder_index(series)
    protocols = build_protocol_summary(series)
    quality = build_quality_summary(series)
    errors = read_errors(connection)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        workbook_guide(database_path).to_excel(writer, sheet_name="Guide", index=False)
        site_summary.to_excel(writer, sheet_name="Site_Summary", index=False)
        write_dataframe_chunks(writer, images, "All_3D_4D")
        protocols.to_excel(writer, sheet_name="Protocol_Matrix", index=False)
        folder_index.to_excel(writer, sheet_name="Folder_Index", index=False)
        quality.to_excel(writer, sheet_name="Quality_Summary", index=False)
        write_dataframe_chunks(writer, errors, "Read_Errors")

    format_workbook(
        workbook_path,
        {"folder_path", "series_folder", "example_file", "file_path"},
    )
    return workbook_path


def build_reports(
    connection: sqlite3.Connection,
    output: Path,
    database_path: Path,
    include_file_sheets: bool,
) -> list[Path]:
    print("\nAggregating DICOM files into series...")
    series = load_series(connection)
    if series.empty:
        raise RuntimeError("The SQLite database contains no DICOM series.")

    csv_path = output / "all_series_inventory.csv"
    series.to_csv(csv_path, index=False, encoding="utf-8-sig")
    outputs = [csv_path]

    sites = sorted(
        series["site_number"].fillna("unknown").astype(str).unique(), key=site_sort_key
    )
    for site in sites:
        site_series = series[
            series["site_number"].fillna("unknown").astype(str) == site
        ].copy()
        print(f"Writing site {site} workbook ({len(site_series):,} series)...")
        outputs.append(
            create_site_workbook(
                site,
                site_series,
                connection,
                output,
                database_path,
                include_file_sheets,
            )
        )

    print("Writing all-sites 3D/4D workbook...")
    outputs.append(create_master_workbook(series, connection, output, database_path))
    return outputs


def site_sort_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value)


def main() -> int:
    args = parse_args()
    root = Path(args.root)
    output = Path(args.output)

    if not args.report_only and not root.exists():
        print(f"ERROR: data root does not exist: {root}", file=sys.stderr)
        return 2
    output.mkdir(parents=True, exist_ok=True)
    database_path = output / "scapis_dicom_inventory.sqlite"
    connection = connect_database(database_path)

    try:
        if not args.report_only:
            print("SCAPIS DICOM inventory")
            print(f"  Data root: {root}")
            print(f"  Output:    {output}")
            print(f"  Workers:   {args.workers}")
            print(f"  Database:  {database_path}")
            print("  Pixel data: never loaded")
            scan_root(
                root=root,
                output=output,
                connection=connection,
                workers=args.workers,
                batch_size=max(1, args.batch_size),
                keep_all_tags=not args.no_all_tags,
                limit=max(0, args.limit),
            )

        outputs = build_reports(
            connection,
            output,
            database_path,
            args.include_file_sheets,
        )
    finally:
        connection.close()

    print("\nCreated:")
    print(f"  {database_path}")
    for path in outputs:
        print(f"  {path}")
    print("\nRe-running the command resumes safely and skips unchanged DICOM files.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
